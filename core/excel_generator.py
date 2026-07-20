# -*- coding: utf-8 -*-
import os

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def generate_reimbursement_excel(files, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "报销单"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ['日期', '费用类型', '金额（元）', '发票明细', '备注']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    ws.freeze_panes = 'A2'

    row_num = 2
    category_rows = {}
    for item in files:
        ws.cell(row=row_num, column=1, value=item.get('date', ''))
        ws.cell(row=row_num, column=2, value=item.get('category', ''))
        ws.cell(row=row_num, column=3, value=float(item.get('amount', 0)))
        link_cell = ws.cell(row=row_num, column=4)
        link_cell.value = "查看发票"
        link_cell.hyperlink = f"E:\\发票\\{os.path.basename(item.get('invoice_path', ''))}"
        link_cell.style = "Hyperlink"
        ws.cell(row=row_num, column=5, value=item.get('remark', ''))
        cat = item.get('category', '')
        category_rows.setdefault(cat, []).append(row_num)
        row_num += 1

    subtotal_rows = []
    for cat, rows in category_rows.items():
        if rows:
            subtotal_row = row_num
            row_num += 1
            subtotal_rows.append(subtotal_row)
            ws.cell(row=subtotal_row, column=2, value=f"小计-{cat}")
            # 直接赋值公式字符串（以=开头）
            ws.cell(row=subtotal_row, column=3,
                    value=f"=SUM({get_column_letter(3)}{rows[0]}:{get_column_letter(3)}{rows[-1]})")
            ws.cell(row=subtotal_row, column=2).fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7",
                                                                   fill_type="solid")
            ws.cell(row=subtotal_row, column=2).font = Font(bold=True)
            ws.cell(row=subtotal_row, column=3).fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7",
                                                                   fill_type="solid")
            ws.cell(row=subtotal_row, column=3).font = Font(bold=True)

    if subtotal_rows:
        total_row = row_num
        ws.cell(row=total_row, column=2, value="合计")
        sum_parts = [f"{get_column_letter(3)}{r}" for r in subtotal_rows]
        ws.cell(row=total_row, column=3, value=f"=SUM({','.join(sum_parts)})")
        total_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        total_font = Font(bold=True, color="FFFFFF", size=12)
        ws.cell(row=total_row, column=2).fill = total_fill
        ws.cell(row=total_row, column=2).font = total_font
        ws.cell(row=total_row, column=3).fill = total_fill
        ws.cell(row=total_row, column=3).font = total_font

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 55
    ws.column_dimensions['E'].width = 15

    wb.save(output_path)
    return output_path


def generate_travel_detail_excel(travel_records, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "交通明细"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    headers = ['日期', '起点', '终点', '距离（km）', '导航截图']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    sorted_records = sorted(travel_records, key=lambda x: x.get('date', ''))
    for idx, rec in enumerate(sorted_records, start=2):
        ws.cell(row=idx, column=1, value=rec.get('date', ''))
        ws.cell(row=idx, column=2, value=rec.get('start', ''))
        ws.cell(row=idx, column=3, value=rec.get('end', ''))
        ws.cell(row=idx, column=4, value=float(rec.get('distance', 0)))
        ws.cell(row=idx, column=5, value=os.path.basename(rec.get('screenshot_path', '')))

    total_row = len(sorted_records) + 2
    ws.cell(row=total_row, column=3, value="合计")
    ws.cell(row=total_row, column=4, value=f"=SUM(D2:D{total_row - 1})")
    ws.cell(row=total_row, column=3).font = Font(bold=True)
    ws.cell(row=total_row, column=4).font = Font(bold=True)

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 40

    wb.save(output_path)
    return output_path
